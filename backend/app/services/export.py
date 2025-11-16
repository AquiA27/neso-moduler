# backend/app/services/export.py
"""
Export Servisi
Raporları Excel ve PDF formatında dışa aktarma
"""
import io
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from ..core.config import settings

logger = logging.getLogger(__name__)


class ExportService:
    """Rapor export servisi"""

    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        title: str = "Rapor",
        sheet_name: str = "Sayfa1",
    ) -> bytes:
        """
        Veriyi Excel dosyasına dönüştür

        Args:
            data: Export edilecek veri (dictionary listesi)
            columns: Kolonlar (None ise tüm kolonlar)
            title: Excel başlığı
            sheet_name: Sayfa adı

        Returns:
            Excel dosyası (bytes)
        """
        # DataFrame oluştur
        if not data:
            df = pd.DataFrame()
        else:
            df = pd.DataFrame(data)
            if columns:
                # Sadece istenen kolonları al
                df = df[columns]

        # Excel writer
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # DataFrame'i yaz
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)

            # Worksheet'i al
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Başlık ekle
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns) if len(df.columns) > 0 else 1)
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = title
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Header stilleri
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Header'lara stil uygula
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=3, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Veri hücrelerine border ekle
            for row in range(4, len(df) + 4):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border

            # Kolon genişliklerini ayarla
            for column in worksheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_to_pdf(
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        title: str = "Rapor",
        subtitle: Optional[str] = None,
        orientation: str = "portrait",
    ) -> bytes:
        """
        Veriyi PDF dosyasına dönüştür

        Args:
            data: Export edilecek veri (dictionary listesi)
            columns: Kolonlar (None ise tüm kolonlar)
            title: PDF başlığı
            subtitle: Alt başlık (opsiyonel)
            orientation: 'portrait' veya 'landscape'

        Returns:
            PDF dosyası (bytes)
        """
        output = io.BytesIO()

        # Sayfa boyutu
        pagesize = landscape(A4) if orientation == "landscape" else A4

        # Document oluştur
        doc = SimpleDocTemplate(
            output,
            pagesize=pagesize,
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
        )

        # Stiller
        styles = getSampleStyleSheet()

        # Başlık stili
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            alignment=TA_CENTER,
            spaceAfter=12,
        )

        # Alt başlık stili
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_CENTER,
            spaceAfter=20,
        )

        # İçerik
        story = []

        # Başlık
        story.append(Paragraph(title, title_style))

        # Alt başlık (tarih bilgisi)
        if subtitle:
            story.append(Paragraph(subtitle, subtitle_style))
        else:
            story.append(Paragraph(f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}", subtitle_style))

        story.append(Spacer(1, 0.5*cm))

        # Veri yoksa
        if not data:
            story.append(Paragraph("Veri bulunamadı.", styles['Normal']))
        else:
            # DataFrame oluştur
            df = pd.DataFrame(data)
            if columns:
                df = df[columns]

            # Tablo verisi hazırla
            table_data = []

            # Header
            table_data.append(list(df.columns))

            # Veriler
            for _, row in df.iterrows():
                table_data.append([str(val) if val is not None else "" for val in row])

            # Tablo oluştur
            table = Table(table_data, repeatRows=1)

            # Tablo stilleri
            table.setStyle(TableStyle([
                # Header stilleri
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

                # Veri stilleri
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),

                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

                # Zebra striping
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))

            story.append(table)

        # PDF oluştur
        doc.build(story)

        output.seek(0)
        return output.getvalue()

    @staticmethod
    def create_daily_report_excel(report_data: Dict[str, Any]) -> bytes:
        """
        Günlük raporu Excel formatında oluştur

        Args:
            report_data: Rapor verisi (sipariş, ödeme, stok vb.)

        Returns:
            Excel dosyası (bytes)
        """
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. Özet Sayfası
            summary_data = {
                "Metrik": [
                    "Toplam Ciro",
                    "Toplam Sipariş",
                    "Ortalama Sepet",
                    "Toplam Gider",
                    "Net Kar",
                ],
                "Değer": [
                    f"{report_data.get('total_revenue', 0):.2f} ₺",
                    report_data.get('total_orders', 0),
                    f"{report_data.get('avg_basket', 0):.2f} ₺",
                    f"{report_data.get('total_expenses', 0):.2f} ₺",
                    f"{report_data.get('net_profit', 0):.2f} ₺",
                ],
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name="Özet", index=False)

            # 2. Siparişler
            if "orders" in report_data and report_data["orders"]:
                df_orders = pd.DataFrame(report_data["orders"])
                df_orders.to_excel(writer, sheet_name="Siparişler", index=False)

            # 3. Ödemeler
            if "payments" in report_data and report_data["payments"]:
                df_payments = pd.DataFrame(report_data["payments"])
                df_payments.to_excel(writer, sheet_name="Ödemeler", index=False)

            # 4. Popüler Ürünler
            if "popular_products" in report_data and report_data["popular_products"]:
                df_products = pd.DataFrame(report_data["popular_products"])
                df_products.to_excel(writer, sheet_name="Popüler Ürünler", index=False)

            # Stillendirme
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Header stilleri
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

                # Kolon genişlikleri
                for column in worksheet.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        output.seek(0)
        return output.getvalue()


# Global export service instance
export_service = ExportService()
